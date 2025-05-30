import datetime as dt
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import SessionLocal, Shift

TEMPLATE = Path(__file__).resolve().parent.parent / "template.xlsx"

def build_excel(period_first: dt.date,
                period_last: dt.date,
                out_path: Path) -> Path:
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    # 列マッピング（名前列＝A, 役職列＝B として C列を1日目扱い）
    date_cols = {period_first + dt.timedelta(days=i): 3 + i
                 for i in range((period_last - period_first).days + 1)}

    db: Session = SessionLocal()
    stmt = select(Shift).where(
        Shift.work_date.between(period_first, period_last)
    )
    shifts = db.scalars(stmt).all()

    row_idx: dict[str, int] = {}
    for sh in shifts:
        r = row_idx.setdefault(sh.staff_name, ws.max_row + 1)
        ws.cell(row=r, column=1, value=sh.staff_name)  # A列
        c = date_cols[sh.work_date]
        cell = ws.cell(row=r, column=c,
                       value=f"{sh.start_time.strftime('%H:%M')}\n{sh.end_time.strftime('%H:%M')}")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    wb.save(out_path)
    return out_path
