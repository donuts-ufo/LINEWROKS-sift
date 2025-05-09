#!/usr/bin/env python3
# shift_export.py
"""
LINE WORKS のグループラインから取得したシフト希望を
「名前 × 1日〜15日の開始・終了時刻」形式で Excel に出力します。

使い方:
    python shift_export.py [--period 前半|後半]

  --period を指定しなければ、
    当日の日付が 1〜5日の場合 → “前半”、
    6日以降             → “後半”
  を自動判定します。

生成ファイル:
    shift_YYYYMM_前半.xlsx  または  shift_YYYYMM_後半.xlsx
"""

import os
import sys
import time
import calendar
import argparse
import requests
import jwt
from datetime import datetime, timedelta
from openpyxl import Workbook

# ─── 設定セクション ────────────────────────────────────
# 必ず自分の環境に合わせて書き換えてください。
CLIENT_ID         = os.getenv("LW_CLIENT_ID",    "SctmsmO3hsWsf9hW3Yrp")
SERVICE_ACCOUNT_ID= os.getenv("LW_SA_ID",        "YOUR_SERVICE_ACCOUNT_ID")
DOMAIN_ID         = os.getenv("LW_DOMAIN_ID",    "YOUR_DOMAIN_ID")
PRIVATE_KEY_PATH  = os.getenv("LW_PRIVATE_KEY",  "path/to/your_private_key.pem")
API_SERVER        = os.getenv("LW_API_SERVER",   "https://apis.worksmobile.com")
CHAT_ID           = os.getenv("LW_CHAT_ID",      "YOUR_CHAT_ID")

TOKEN_URL         = "https://auth.worksmobile.com/oauth2/v2.0/token"
MESSAGE_URL_TPL   = f"{API_SERVER}/{DOMAIN_ID}/conversation/v1/spaces/{{chat_id}}/messages"
# ────────────────────────────────────────────────────────


def get_access_token():
    """JWT を生成して OAuth2 トークンを取得"""
    with open(PRIVATE_KEY_PATH, "r") as f:
        private_key = f.read()
    now = int(time.time())
    payload = {
        "iss": CLIENT_ID,
        "sub": SERVICE_ACCOUNT_ID,
        "iat": now,
        "exp": now + 3600,
        "domain": DOMAIN_ID
    }
    headers = {"alg": "RS256", "typ": "JWT", "kid": SERVICE_ACCOUNT_ID}
    assertion = jwt.encode(payload, private_key, algorithm="RS256", headers=headers)

    data = {
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
        "assertion": assertion
    }
    resp = requests.post(TOKEN_URL, data=data)
    resp.raise_for_status()
    return resp.json()["access_token"]


def get_messages(token, start_time, end_time):
    """指定期間のメッセージを取得"""
    url = MESSAGE_URL_TPL.format(chat_id=CHAT_ID)
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "startTime": start_time.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z",
        "endTime":   end_time.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z",
        "limit":     100
    }
    resp = requests.get(url, headers=headers, params=params)
    resp.raise_for_status()
    return resp.json().get("items", [])


def parse_shifts(messages):
    """
    メッセージ本文から
      名前：○○
      1日 09:00 18:00
      …
    を正規表現で抜き出し、
    { "山田太郎": {1:{"start":"09:00","end":"18:00"}, … }, … }
    の構造に整形します。
    """
    import re
    from collections import defaultdict

    re_name = re.compile(r"名前：(.+)")
    re_line = re.compile(r"(\d{1,2})日\s+(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})")

    results = defaultdict(dict)
    for item in messages:
        body = item.get("body", "")
        m_name = re_name.search(body)
        if not m_name:
            continue
        name = m_name.group(1).strip()

        for line in body.splitlines():
            m = re_line.match(line.strip())
            if m:
                day = int(m.group(1))
                results[name][day] = {"start": m.group(2), "end": m.group(3)}
    return results


def write_excel(data, period):
    """
    openpyxl で Excel ワークブックを作成。
    1行目に「名前 | 1日開始 | 1日終了 | … | 15日開始 | 15日終了」
    2行目以降に各スタッフのデータを書き込み、
    shift_YYYYMM_<period>.xlsx として保存。
    """
    ym    = datetime.now().strftime("%Y%m")
    fname = f"shift_{ym}_{period}.xlsx"
    wb    = Workbook()
    ws    = wb.active
    ws.title = "シフト"

    # ヘッダ
    ws.cell(row=1, column=1, value="名前")
    for d in range(1, 16):
        ws.cell(row=1, column=2*d,   value=f"{d}日 開始")
        ws.cell(row=1, column=2*d+1, value=f"{d}日 終了")

    # データ部
    row = 2
    for name, shifts in data.items():
        ws.cell(row=row, column=1, value=name)
        for d in range(1, 16):
            sc = 2*d
            ec = 2*d + 1
            s = shifts.get(d, {}).get("start", "")
            e = shifts.get(d, {}).get("end", "")
            ws.cell(row=row, column=sc, value=s)
            ws.cell(row=row, column=ec, value=e)
        row += 1

    wb.save(fname)
    return fname


def get_period_window(period):
    """
    ── 前半 ──
      例：2025年5月の場合
      start = 2025/04/20 00:00:00
      end   = 2025/05/05 23:59:59

    ── 後半 ──
      start = 今月1日 00:00:00
      end   = 今月最終日 23:59:59
    """
    now = datetime.now()
    if period == "前半":
        prev = (now.replace(day=1) - timedelta(days=1))
        start = prev.replace(day=20, hour=0, minute=0, second=0)
        end   = now.replace(day=5, hour=23, minute=59, second=59)
    else:
        last = calendar.monthrange(now.year, now.month)[1]
        start = now.replace(day=1, hour=0, minute=0, second=0)
        end   = now.replace(day=last, hour=23, minute=59, second=59)
    return start, end


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--period", choices=["前半", "後半"],
                   help="手動で“前半”か“後半”を指定（省略可）")
    args = p.parse_args()

    # period の決定
    if args.period:
        period = args.period
    else:
        period = "前半" if datetime.now().day <= 5 else "後半"

    print(f"→ 対象期間: {period}")
    start, end = get_period_window(period)
    print(f"→ 取得期間: {start} 〜 {end}")

    # 1) トークン取得
    token = get_access_token()
    # 2) メッセージ取得
    msgs  = get_messages(token, start, end)
    # 3) パース
    data  = parse_shifts(msgs)
    # 4) Excel 出力
    fname = write_excel(data, period)
    print(f"✅ Excel を出力しました → {fname}")


if __name__ == "__main__":
    main()
